"""
核心功能说明：
1. 源表名强制去除"syt_"前缀（不区分大小写）
2. 目标前缀硬编码为"tb_xg_yx_"
3. 仅对匹配成功的源表名标注绿色
4. 保留原始Excel数据不变
"""

import re
import pandas as pd
import openpyxl
from openpyxl.styles  import PatternFill
from sqlalchemy import create_engine, text
from typing import List, Dict


class TableMatcher:
    def __init__(self, excel_path: str, db_config: Dict[str, str]):
        self.excel_path  = excel_path
        self.db_config  = db_config
        self.target_prefix  = "tb_xg_xggl_"
        self.matched_source  = []  # 存储原始匹配表名

        # 高亮样式配置
        self.green_highlight  = PatternFill(
            start_color='92D050',
            fill_type='solid'
        )

        # 初始化数据库连接
        self.engine  = create_engine(
            f"mysql+pymysql://{db_config['user']}:{db_config['password']}"
            f"@{db_config['host']}:{db_config['port']}/{db_config['database']}"
            "?charset=utf8mb4"
        )

    def _transform_name(self, raw_name: str) -> str:
        """执行表名转换三步流程"""
        # 步骤1：去除前缀
        step1 = re.sub(r'^ST_',  '', raw_name, flags=re.IGNORECASE)
        # 步骤2：过滤非法字符
        step2 = re.sub(r'[^a-zA-Z0-9_]',  '', step1.strip())
        # 步骤3：标准化格式
        return step2.lower().replace('  ', '_')

    def _read_source_names(self) -> List[str]:
        """读取Excel源数据"""
        df = pd.read_excel(self.excel_path)
        target_col = next(col for col in df.columns  if "源系统表英文名称(*)" in str(col))
        return df[target_col].dropna().astype(str).unique().tolist()

    def _fetch_target_tables(self) -> List[str]:
        """获取数据库目标表列表"""
        with self.engine.connect()  as conn:
            result = conn.execute(text(
                f"SELECT LOWER(table_name) FROM information_schema.tables  "
                f"WHERE table_name LIKE '{self.target_prefix}%'  "
                "AND table_schema = DATABASE()"
            ))
            return [row[0] for row in result]

    def _build_mapping(self, source_list: List[str]) -> Dict[str, str]:
        """构建映射关系字典"""
        return {
            src: f"{self.target_prefix}{self._transform_name(src)}"
            for src in source_list
        }

    def _apply_highlight(self):
        """执行绿色标注逻辑"""
        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb.active

        # 定位目标列
        target_col = next(
            col.column_letter  for col in ws[1]
            if "源系统表英文名称(*)" in str(col.value)
        )

        # 仅标注匹配项
        for row in ws.iter_rows(min_row=2):
            cell = row[ord(target_col) - 65]  # 列字母转ASCII码
            if cell.value  in self.matched_source:
                cell.fill  = self.green_highlight

        output_path = self.excel_path.replace('.xlsx',  '_MATCHED.xlsx')
        wb.save(output_path)
        return output_path

    def execute(self) -> dict:
        """主执行流程"""
        # 数据准备
        source_names = self._read_source_names()
        target_tables = self._fetch_target_tables()

        # 匹配分析
        mapping = self._build_mapping(source_names)
        self.matched_source  = [
            src for src, target in mapping.items()
            if target in target_tables
        ]

        # 生成结果
        result_file = self._apply_highlight()

        return {
            "total": len(source_names),
            "matched": len(self.matched_source),
            "match_ratio": f"{len(self.matched_source)/len(source_names):.0%}",
            "output_file": result_file,
            "demo": [f"{k} → {v}" for k, v in list(mapping.items())[:3]]
        }


# 使用示例
if __name__ == "__main__":
    # 数据库配置
    DB_CONF = {
        "host": "10.10.10.215",
        "user": "etladmin",
        "password": "etladmin123",
        "database": "HG_SourceDB",
        "port": 3306
    }
    EXCEL_PATH = "source_excel/12.xlsx"

    # 执行匹配
    processor = TableMatcher(EXCEL_PATH, DB_CONF)
    report = processor.execute()

    # 输出报告
    print(f"✅ 匹配完成：{report['matched']}/{report['total']}")
    print(f"📈 匹配率：{report['match_ratio']}")
    print(f"💡 样例映射：")
    for example in report['demo']:
        print(f"  ▪ {example}")