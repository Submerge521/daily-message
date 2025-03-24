"""
精准前缀表名比对系统 (v2.1)
更新时间：2025-03-13 10:44
核心特性：
1. 固定前缀硬编码匹配（tb_czyth_gzc_）
2. 智能大小写转换与字符清洗
3. 特殊符号过滤机制
4. 实时匹配结果可视化
"""

import re
import pandas as pd
import openpyxl
from openpyxl.styles  import PatternFill
from sqlalchemy import create_engine, text
from typing import List, Dict


class FixedPrefixComparator:
    def __init__(self, excel_path: str, db_config: Dict[str, str]):
        self.excel_path  = excel_path
        self.db_config  = db_config
        self.fixed_prefix  = "tb_kjc_zw_"  # 硬编码指定前缀
        self.matched_tables  = []

        # 样式配置
        self.highlight_style  = PatternFill(
            start_color='92D050',
            fill_type='solid'
        )

        # 初始化数据库连接
        self.engine  = create_engine(
            f"mysql+pymysql://{db_config['user']}:{db_config['password']}"
            f"@{db_config['host']}:{db_config['port']}/{db_config['database']}"
            "?charset=utf8mb4"
        )

    def _clean_table_name(self, raw_name: str) -> str:
        """
        执行表名标准化处理：
        1. 移除首尾空格
        2. 替换内部空格为下划线
        3. 过滤非法字符
        4. 统一小写化
        """
        # 替换连续空格为单个下划线
        step1 = re.sub(r'\s+',  '_', raw_name.strip())
        # 过滤非字母数字下划线的字符
        step2 = re.sub(r'[^a-zA-Z0-9_]',  '', step1)
        # 转换为小写
        return step2.lower()

    def _get_excel_names(self) -> List[str]:
        """读取Excel源数据并去重"""
        df = pd.read_excel(self.excel_path)
        target_col = next(col for col in df.columns  if "源系统表英文名称(*)" in str(col))
        return df[target_col].dropna().astype(str).unique().tolist()

    def _get_mysql_tables(self) -> List[str]:
        """获取目标前缀的数据库表"""
        with self.engine.connect()  as conn:
            result = conn.execute(text(
                "SELECT table_name FROM information_schema.tables  "
                f"WHERE table_name LIKE '{self.fixed_prefix}%'  "
                "AND table_schema = DATABASE()"
            ))
            return [row[0] for row in result]

    def _generate_mapping(self, excel_names: List[str]) -> Dict[str, str]:
        """生成Excel表名到数据库表名的映射"""
        return {
            orig: f"{self.fixed_prefix}{self._clean_table_name(orig)}"
            for orig in excel_names
        }

    def _highlight_excel(self):
        """在Excel中标注匹配项"""
        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb.active

        # 定位目标列
        target_col = next(
            col.column_letter  for col in ws[1]
            if "源系统表英文名称(*)" in str(col.value)
        )

        # 遍历标注
        for row in ws.iter_rows(min_row=2):
            cell = row[ord(target_col) - 65]  # A=65
            if cell.value  in self.matched_tables:
                cell.fill  = self.highlight_style

        output_path = self.excel_path.replace('.xlsx',  '_RESULTS.xlsx')
        wb.save(output_path)
        return output_path

    def execute(self) -> dict:
        """执行比对全流程"""
        # 数据获取
        excel_names = self._get_excel_names()
        mysql_tables = self._get_mysql_tables()

        # 映射匹配
        mapping = self._generate_mapping(excel_names)
        self.matched_tables  = [k for k, v in mapping.items()  if v in mysql_tables]

        # 生成结果
        result_file = self._highlight_excel()

        return {
            "total": len(excel_names),
            "matched": len(self.matched_tables),
            "match_rate": f"{len(self.matched_tables)/len(excel_names):.0%}",
            "output_path": result_file,
            "sample": [
                f"{excel} → {mapping[excel]}"
                for excel in self.matched_tables[:3]
            ]
        }


# 使用示例
if __name__ == "__main__":
    # 配置参数
    DB_CONFIG = {
        "host": "10.10.10.215",
        "user": "etladmin",
        "password": "etladmin123",
        "database": "HG_SourceDB",
        "port": 3306
    }
    # EXCEL_PATH = "source_excel/16.xlsx"
    EXCEL_PATH = "compare/2_preprocessed_20250314_1211.xlsx"
    # 执行比对
    comparator = FixedPrefixComparator(EXCEL_PATH, DB_CONFIG)
    report = comparator.execute()

    # 打印结果报告
    print(f"📊 匹配完成报告（{report['matched']}/{report['total']}）")
    print(f"📈 匹配率: {report['match_rate']}")
    print(f"💾 结果文件: {report['output_path']}")
    print("🔍 样例匹配:")
    for sample in report['sample']:
        print(f"  └ {sample}")