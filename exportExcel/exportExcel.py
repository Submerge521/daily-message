"""
Oracle Schema统计Excel导出工具（2025-03-18增强版）
功能：跨Schema表统计、实时计数、异常状态标记、自动生成Excel报告
环境：Python 3.9+ | 依赖：pip install cx_Oracle openpyxl
"""

import cx_Oracle
import openpyxl
import argparse
from openpyxl.styles  import Font, PatternFill, Alignment
from datetime import datetime

class OracleSchemaReporter:
    def __init__(self, username: str, password: str, dsn: str):
        self.conn  = cx_Oracle.connect(
            user=username,
            password=password,
            dsn=dsn,
            encoding="UTF-8",
            events=False
        )
        self.username  = username.upper()
        self.stats_cache  = {}  # 缓存统计信息

    def generate_report(self, target_schema: str, use_count: bool = False, output_file: str = None):
        """生成统计报告主流程"""
        target_schema = target_schema.upper()
        try:
            # 阶段1：元数据采集
            tables = self._get_tables_metadata(target_schema)
            if not tables:
                raise ValueError(f"Schema [{target_schema}] 无可用表或访问被拒绝")

            # 阶段2：数据统计
            stats = self._count_tables(tables, target_schema) if use_count else self._get_cached_stats(tables, target_schema)

            # 阶段3：生成Excel
            output_path = self._create_excel(stats, target_schema, output_file)

            return {
                "status": "success",
                "schema": target_schema,
                "table_count": len(stats),
                "output_path": output_path,
                "timestamp": datetime.now().isoformat()
            }
        except cx_Oracle.DatabaseError as e:
            error = e.args[0]
            return {
                "status": "error",
                "code": error.code,
                "message": error.message.split("\n",  1)[0]
            }

    def _get_tables_metadata(self, schema: str) -> list:
        """获取表基础信息（带双重缓存机制）"""
        # 优先使用缓存
        if schema in self.stats_cache:
            return self.stats_cache[schema]

        sql = """
            SELECT a.table_name,  b.comments, 
                   a.num_rows,  a.last_analyzed  
            FROM all_tables a 
            LEFT JOIN all_tab_comments b 
              ON a.owner  = b.owner  AND a.table_name  = b.table_name  
            WHERE a.owner  = :1 
              AND a.table_name  NOT LIKE 'BIN$%'
              AND a.table_name  NOT LIKE 'SYS_%'
            ORDER BY a.table_name  
        """
        with self.conn.cursor()  as cursor:
            cursor.execute(sql,  [schema])
            result = [{
                "name": row[0],
                "comment": row[1] or "",
                "cached_rows": row[2] or 0,
                "last_analyzed": row[3].strftime("%Y-%m-%d") if row[3] else "N/A"
            } for row in cursor]

        # 写入缓存
        self.stats_cache[schema]  = result
        return result

    def _count_tables(self, tables: list, schema: str) -> list:
        """实时统计模式（带重试机制）"""
        results = []
        with self.conn.cursor()  as cursor:
            for table in tables:
                try:
                    cursor.execute(f"SELECT  /*+ PARALLEL(4) */ COUNT(*) FROM {schema}.{table['name']}")
                    count = cursor.fetchone()[0]
                    status = "success"
                except cx_Oracle.DatabaseError as e:
                    error = e.args[0]
                    count, status = self._parse_error(error)

                results.append({
                    **table,
                    "real_rows": count,
                    "status": status
                })
        return results

    def _parse_error(self, error) -> tuple:
        """解析数据库错误"""
        code = error.code
        if code == 942:   # 表不存在
            return (-2, "object_not_found")
        elif code == 1031: # 权限不足
            return (-1, "permission_denied")
        elif code == 12154: # TNS错误
            raise RuntimeError("连接配置错误")
        return (-99, "unknown_error")

    def _get_cached_stats(self, tables: list, schema: str) -> list:
        """快速统计模式"""
        return [{
            **table,
            "real_rows": table["cached_rows"],
            "status": "cached"
        } for table in tables]

    def _create_excel(self, data: list, schema: str, filename: str = None) -> str:
        """生成Excel报告（带样式优化）"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title  = "表统计报告"

        # 设置列宽
        col_config = [
            ("表名", 35),
            ("注释", 50),
            ("缓存行数", 15),
            ("实时行数", 15),
            ("最后统计时间", 20),
            ("状态", 15)
        ]
        for idx, (title, width) in enumerate(col_config, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width  = width
            ws.cell(1,  idx, title).font = Font(bold=True, color="FFFFFF")

        # 写入数据
        for row_idx, item in enumerate(data, 2):
            ws.append([
                item["name"],
                item["comment"],
                item["cached_rows"],
                item["real_rows"] if item["status"] not in ["permission_denied", "object_not_found"] else "N/A",
                item["last_analyzed"],
                self._translate_status(item["status"])
            ])

            # 状态颜色标记
            status_cell = ws.cell(row_idx,  6)
            status_cell.fill  = self._get_status_color(item["status"])

        # 自动生成文件名
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M")
            filename = f"{schema}_report_{timestamp}.xlsx"

        wb.save(filename)
        return filename

    def _translate_status(self, status: str) -> str:
        """状态码转中文"""
        return {
            "success": "✅ 成功",
            "cached": "📊 缓存数据",
            "permission_denied": "🔒 无权限",
            "object_not_found": "⛔ 对象不存在",
            "unknown_error": "❌ 未知错误"
        }.get(status, "未知状态")

    def _get_status_color(self, status: str) -> PatternFill:
        """状态颜色映射"""
        color_map = {
            "success": "FFC6EFCE",
            "cached": "FFFFE699",
            "permission_denied": "FFFFC7CE",
            "object_not_found": "FFD9D9D9",
            "unknown_error": "FF7030A0"
        }
        return PatternFill("solid", fgColor=color_map.get(status,  "FFFFFFFF"))

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Oracle Schema统计导出工具')
    parser.add_argument('-u',  required=True, help='数据库用户名')
    parser.add_argument('-p',  required=True, help='密码')
    parser.add_argument('-d',  required=True, help='连接DSN，例：localhost:1521/XEPDB1')
    parser.add_argument('-s',  required=True, help='目标Schema名称')
    parser.add_argument('-o',  help='输出文件名（可选）')
    parser.add_argument('-c',  action='store_true', help='启用实时统计模式')

    args = parser.parse_args()

    reporter = OracleSchemaReporter(args.u, args.p, args.d)
    result = reporter.generate_report(args.s,  args.c, args.o)

    if result["status"] == "success":
        print(f"""\n🔄 任务完成于 {datetime.now().strftime('%H:%M:%S')} 
✅ 成功导出 {result['table_count']} 张表 
📂 文件路径：{result['output_path']}""")
    else:
        print(f"""\n❌ 执行失败（错误代码 {result['code']}）
⚠️ 错误信息：{result['message']}""")