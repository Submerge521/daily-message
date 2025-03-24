"""
数据字典生成工具v1.6（双列合并完整版）
最后更新：2025-03-19 11:56
"""
import configparser
import os
from sqlalchemy import create_engine, inspect
import pandas as pd
import openpyxl
from openpyxl.styles  import Font, Alignment, Border, Side
from datetime import datetime
from collections import defaultdict

class DataDictGenerator:
    def __init__(self):
        self.config  = self._load_config()
        self.engine  = self._create_engine()
        self._validate_config()
        self.table_rows  = defaultdict(list)  # 记录每张表的Excel行号

    # █████████████████████ 配置处理 █████████████████████
    def _load_config(self):
        """加载并净化配置文件"""
        config = configparser.ConfigParser(interpolation=None)
        config.read('config.ini',  encoding='utf-8')
        # 移除注释内容（支持#和;两种注释符）
        for section in config.sections():
            for key in config[section]:
                raw_value = config[section][key]
                cleaned_value = raw_value.split('#')[0].split(';')[0].strip()
                config[section][key] = cleaned_value
        return config

    def _validate_config(self):
        """配置完整性校验"""
        required_keys = {
            'database': ['db_type', 'host', 'port', 'username', 'password', 'database'],
            'export': ['output_file', 'sheet_name', 'max_column_width']
        }
        for section, keys in required_keys.items():
            missing = [key for key in keys if not self.config[section].get(key)]
            if missing:
                raise ValueError(f"配置缺失：[{section}] {', '.join(missing)}")

    # █████████████████████ 数据库连接 █████████████████████
    def _create_engine(self):
        """创建跨数据库引擎"""
        db_type = self.config['database']['db_type'].lower()
        connectors = {
            'mysql': lambda: f"mysql+pymysql://{self.config['database']['username']}:{self.config['database']['password']}@{self.config['database']['host']}:{self.config['database']['port']}/{self.config['database']['database']}",
            'postgresql': lambda: f"postgresql+psycopg2://{self.config['database']['username']}:{self.config['database']['password']}@{self.config['database']['host']}:{self.config['database']['port']}/{self.config['database']['database']}",
            'oracle': lambda: f"oracle+cx_oracle://{self.config['database']['username']}:{self.config['database']['password']}@{self.config['database']['host']}:{self.config['database']['port']}/?service_name={self.config['database']['database']}",
            'sqlserver': lambda: f"mssql+pyodbc://{self.config['database']['username']}:{self.config['database']['password']}@{self.config['database']['host']}:{self.config['database']['port']}/{self.config['database']['database']}?driver=ODBC+Driver+17+for+SQL+Server"
        }
        return create_engine(connectors[db_type]())

    # █████████████████████ 数据采集 █████████████████████
    def _get_table_comment(self, inspector, table_name):
        """跨数据库表注释提取"""
        try:
            comment = inspector.get_table_comment(table_name)
            if isinstance(comment, dict):  # PostgreSQL处理
                return comment.get('text',  '')
            elif isinstance(comment, tuple):  # MySQL处理
                return comment[0] or ''
            return str(comment) if comment else ''
        except Exception as e:
            print(f"[WARN] 表注释获取失败：{table_name} - {str(e)}")
            return ''

    def _get_metadata(self):
        """获取完整元数据并记录行号"""
        inspector = inspect(self.engine)
        meta_data = []

        for table_name in inspector.get_table_names():
            table_comment = self._get_table_comment(inspector, table_name)
            columns = inspector.get_columns(table_name)

            for col in columns:
                # 数据类型标准化处理（移除长度信息）
                data_type = str(col['type']).split('(')[0].split(' ')[0]
                default_value = str(col['default']) if col['default'] is not None else ''

                meta_data.append({
                    '表名称': table_name,
                    '字段名称': col['name'],
                    '数据类型': data_type.upper(),
                    '可为空': '是' if col['nullable'] else '否',
                    '默认值': default_value,
                    '字段注释': col.get('comment',  ''),
                    '表注释': table_comment if table_comment else '（无注释）'
                })
                # 记录当前数据行号（Excel从第2行开始）
                self.table_rows[table_name].append(len(meta_data)  + 1)

        return pd.DataFrame(meta_data)

    # █████████████████████ 样式处理 █████████████████████
    def _style_excel(self, writer):
        """双列合并核心逻辑"""
        workbook = writer.book
        worksheet = workbook[self.config['export']['sheet_name']]

        # 通用样式定义
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        header_font = Font(name='微软雅黑', bold=True, color="FFFFFF")
        header_fill = openpyxl.styles.PatternFill(fgColor="2E75B5",  fill_type="solid")

        # 合并表名称（A列）和表注释（G列）
        for table_name, rows in self.table_rows.items():
            if len(rows) > 1:
                start_row, end_row = rows[0], rows[-1]
                # 合并表名称单元格（第1列）
                worksheet.merge_cells(start_row=start_row,  end_row=end_row, start_column=1, end_column=1)
                cell_name = worksheet.cell(row=start_row,  column=1)
                cell_name.alignment  = center_alignment
                cell_name.border  = thin_border
                # 合并表注释单元格（第7列）
                worksheet.merge_cells(start_row=start_row,  end_row=end_row, start_column=7, end_column=7)
                cell_comment = worksheet.cell(row=start_row,  column=7)
                cell_comment.alignment  = center_alignment
                cell_comment.border  = thin_border

        # 标题样式设置
        for col in range(1, 8):  # 列A到列G
            cell = worksheet.cell(row=1,  column=col)
            cell.font  = header_font
            cell.fill  = header_fill
            cell.alignment  = center_alignment
            cell.border  = thin_border

        # 动态列宽设置（限制最大宽度）
        column_widths = {
            'A': 25,  # 表名称
            'B': 20,  # 字段名称
            'C': 15,  # 数据类型
            'D': 10,  # 可为空
            'E': 18,  # 默认值
            'F': 30,  # 字段注释
            'G': 40   # 表注释
        }
        for col_letter, width in column_widths.items():
            worksheet.column_dimensions[col_letter].width  = width

        # 冻结首行、添加筛选
        worksheet.freeze_panes  = 'A2'
        worksheet.auto_filter.ref  = worksheet.dimensions

    # █████████████████████ 主流程 █████████████████████
    def generate(self):
        """生成主流程"""
        try:
            df = self._get_metadata()
            if df.empty:
                raise ValueError("未获取到任何表结构信息，请检查数据库连接和权限")

            # 生成带精确时间戳的文件名
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            raw_filename = self.config['export']['output_file'].replace('%TIMESTAMP%',  timestamp)
            filename = os.path.abspath(raw_filename)

            # 自动创建多级目录
            output_dir = os.path.dirname(filename)
            if output_dir and not os.path.exists(output_dir):
                os.makedirs(output_dir,  exist_ok=True)

            # 写入Excel并应用样式
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer,  index=False, sheet_name=self.config['export']['sheet_name'])
                self._style_excel(writer)

            print(f"[{datetime.now().strftime('%Y-%m-%d  %H:%M:%S')}] 文件已生成：{filename}")
            return True
        except Exception as e:
            print(f"[ERROR] 生成失败：{str(e)}")
            return False

if __name__ == "__main__":
    DataDictGenerator().generate()