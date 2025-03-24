import cx_Oracle
import configparser
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


def read_config():
    """读取配置文件"""
    config = configparser.ConfigParser()
    config.read('config.ini', encoding='utf-8')
    return config['oracle']


def get_db_connection(config):
    """建立数据库连接"""
    dsn = cx_Oracle.makedsn(config['host'], config['port'],
                            service_name=config['service_name'])
    return cx_Oracle.connect(config['user'], config['password'], dsn)


def query_metadata(conn, schema):
    """获取元数据"""
    with conn.cursor() as cursor:
        # 获取表清单及注释
        cursor.execute(""" 
            SELECT table_name, comments 
            FROM all_tab_comments 
            WHERE owner = :owner AND table_type = 'TABLE'  
            ORDER BY table_name""", [schema.upper()])
        tables = cursor.fetchall()

        # 获取字段详细信息
        metadata = []
        for table, tab_comment in tables:
            cursor.execute(""" 
                SELECT 
                    col.column_name, 
                    col.data_type  || 
                        CASE 
                            WHEN col.data_type  = 'NUMBER' AND col.data_precision  IS NOT NULL 
                                THEN '(' || col.data_precision  || 
                                     CASE WHEN col.data_scale  > 0 THEN ',' || col.data_scale  ELSE '' END || ')'
                            WHEN col.data_type  IN ('VARCHAR2', 'CHAR', 'NVARCHAR2', 'NCHAR') 
                                THEN '(' || col.data_length  || ')'
                            ELSE ''
                        END as data_type,
                    col.nullable, 
                    col.data_default, 
                    com.comments  
                FROM all_tab_columns col 
                LEFT JOIN all_col_comments com 
                    ON col.owner  = com.owner  
                    AND col.table_name  = com.table_name  
                    AND col.column_name  = com.column_name  
                WHERE col.owner  = :owner AND col.table_name  = :table_name 
                ORDER BY col.column_id""",
                           [schema.upper(), table])
            columns = cursor.fetchall()
            metadata.append((table, tab_comment, columns))
        return metadata


def generate_excel(metadata, filename):
    """生成带合并单元格的Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "数据字典"

    # 设置表头
    headers = ['表名', '表注释', '字段名', '数据类型', '允许为空', '默认值', '字段注释']
    ws.append(headers)

    # 写入数据并记录合并区域
    current_row = 2
    merge_ranges = []

    for table, comment, columns in metadata:
        start_row = current_row
        for i, col in enumerate(columns):
            row_data = [
                table if i == 0 else None,
                comment if i == 0 else None,
                col[0], col[1],
                'Y' if col[2] == 'Y' else 'N',
                str(col[3]).strip() if col[3] else '',
                col[4] or ''
            ]
            ws.append(row_data)
            current_row += 1

            # 记录合并范围
        end_row = current_row - 1
        merge_ranges.append(f'A{start_row}:A{end_row}')
        merge_ranges.append(f'B{start_row}:B{end_row}')

        # 执行合并操作
    for range in merge_ranges:
        ws.merge_cells(range)

        # 设置对齐格式
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical='center', wrap_text=True)

    # 调整列宽
    col_widths = [20, 30, 20, 15, 10, 15, 30]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    wb.save(filename)


if __name__ == "__main__":
    config = read_config()
    conn = get_db_connection(config)
    metadata = query_metadata(conn, config['schema'])
    generate_excel(metadata, config['output_file'])
    print(f"数据字典已生成至：{config['output_file']}")