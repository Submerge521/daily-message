import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


def process_excel(table2_path, output_path):
    try:
        # 1. 数据校验与读取
        required_columns = ['表名', '字段名', '字段注释']
        df_dict = pd.read_excel(table2_path)

        # 校验必要列存在性
        missing_cols = set(required_columns) - set(df_dict.columns)
        if missing_cols:
            raise ValueError(f"输入文件缺失关键列：{', '.join(missing_cols)}")

        # 2. 数据结构构建
        processed_data = []
        table_groups = {}  # 记录表名合并范围 {表名: (起始行, 结束行)}
        current_table = None

        for idx, (_, row) in enumerate(df_dict.iterrows(), 1):
            # 生成ETL任务路径（科技处/知网）
            etl_task = f"自助ETL\\数据源抽取toMysql集群\\科技处\\知网\\{row['表名']}"

            # 记录表名合并范围
            if row['表名'] != current_table:
                current_table = row['表名']
                start_row = idx + 2  # 表头占2行
                table_groups[current_table] = {'start': start_row, 'end': start_row}
            else:
                table_groups[current_table]['end'] = idx + 2

                # 构建数据记录
            processed_data.append({
                "序号": idx,
                "来源表名": row['表名'],
                "来源字段": row['字段名'],
                "字段注释": row['字段注释'],
                "入仓表名": f"tb_kjc_zw_{row['表名']}",
                "ETL任务路径": etl_task
            })

        # 3. Excel文件生成
        wb = Workbook()
        ws = wb.active

        # 3.1 构建多级表头
        headers = [
            ["序号", "数据来源信息", "", "", "", "", "数据采集方式", "", "数据进入数仓后的相关信息", "", "", "", "", "",
             ""],
            ["", "数据来源部门", "数据来源视图名/表名", "数据来源字段名", "来源字段中文含义", "数据是否采集",
             "采集方式", "ETL任务位置与任务名称", "入仓后的数据表", "入仓后的字段",
             "有无数据", "是否已通过可视化展示", "备注", "数据量", "ETL定时情况"]
        ]

        # 写入表头并设置样式
        bold_font = Font(bold=True)
        for row in headers:
            ws.append(row)
            for cell in ws[ws.max_row]:
                cell.font = bold_font

                # 合并表头区域
        merge_ranges = [
            ('A1:A2', "序号"),
            ('B1:F1', "数据来源信息"),
            ('G1:H1', "数据采集方式"),
            ('I1:O1', "数据入仓信息")
        ]
        for range_str, _ in merge_ranges:
            ws.merge_cells(range_str)

            # 3.2 写入数据内容
        for data_row in dataframe_to_rows(pd.DataFrame(processed_data), index=False, header=False):
            ws.append([
                data_row[0],  # A列：序号
                "",  # B列：数据来源部门
                data_row[1],  # C列：来源表名
                data_row[2],  # D列：来源字段名
                data_row[3],  # E列：字段注释
                "√",  # F列：数据是否采集（固定值）
                "ETL采集",  # G列：采集方式（固定值）
                data_row[5],  # H列：ETL任务路径
                data_row[4],  # I列：入仓表名
                data_row[2],  # J列：入仓字段
                *[""] * 3,  # K-M列：空字段
                "每天晚上2点执行"  # N列：ETL定时情况
            ])

        # 3.3 合并相同表名的单元格
        for table_name, ranges in table_groups.items():
            if ranges['end'] > ranges['start']:
                # 合并C列（来源表名）
                ws.merge_cells(
                    start_row=ranges['start'],
                    end_row=ranges['end'],
                    start_column=3,
                    end_column=3
                )
                # 合并H列（ETL任务路径）
                ws.merge_cells(
                    start_row=ranges['start'],
                    end_row=ranges['end'],
                    start_column=8,
                    end_column=8
                )
                # 合并I列（入仓表名）
                ws.merge_cells(
                    start_row=ranges['start'],
                    end_row=ranges['end'],
                    start_column=9,
                    end_column=9
                )

        # 3.4 设置列宽
        col_width_config = {
            3: 25,  # C列（来源表名）
            8: 40,  # H列（ETL任务路径）
            9: 30,  # I列（入仓表名）
            14: 18  # N列（ETL定时情况）
        }
        for col_idx in range(1, 15):
            column_letter = get_column_letter(col_idx)
            if col_idx in col_width_config:
                ws.column_dimensions[column_letter].width = col_width_config[col_idx]
            else:
                ws.column_dimensions[column_letter].width = 15

                # 4. 保存与验证
        wb.save(output_path)
        print(f"✅ 文件生成成功：{output_path}")
        print(f"▸ 合并表名数量：{len(table_groups)} 个")
        print(f"▸ 示例入仓表名：tb_kjc_zw_{df_dict['表名'].iloc[0]}")

    except Exception as e:
        print(f"❌ 运行错误：{str(e)}")
        if "Permission denied" in str(e):
            print("→ 请关闭已打开的Excel文件后重试")


# 执行示例
if __name__ == "__main__":
    process_excel(
        table2_path="表4.xlsx",
        output_path="数据映射表_20250324_ZW.xlsx"
    )