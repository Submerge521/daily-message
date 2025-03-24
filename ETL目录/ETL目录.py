import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


def process_excel(table2_path, output_path):
    try:
        # 1. 数据校验与读取
        required_columns = ['表名称', '字段名称', '字段注释', '表注释']
        df_dict = pd.read_excel(table2_path)

        # 校验必要列存在性
        missing_cols = set(required_columns) - set(df_dict.columns)
        if missing_cols:
            raise ValueError(f"输入文件缺失关键列：{', '.join(missing_cols)}")

        # 2. 数据结构构建
        processed_data = []
        table_groups = {}  # 记录表名合并范围 {表名: (起始行, 结束行)}
        current_table = None
        for idx, (_, row) in enumerate(df_dict.iterrows(), 1):
            # 生成复合表名和ETL任务路径
            table_full_name = f"{row['表名称']}--{row['表注释']}"
            etl_task = f"自助ETL\\数据源抽取toMysql集群\\体育部\\场馆预约\\{table_full_name}"

            # 记录表名合并范围
            if row['表名称'] != current_table:
                current_table = row['表名称']
                start_row = idx + 2  # +2补偿表头占用的2行
                table_groups[current_table] = {'start': start_row, 'end': start_row}
            else:
                table_groups[current_table]['end'] = idx + 2

                # 构建数据记录
            processed_data.append({
                "序号": idx,
                "来源表名": table_full_name,
                "来源字段": row['字段名称'],
                "字段注释": row['字段注释'],
                "入仓表名": f"tb_tyb_cgyy_{row['表名称']}",
                "ETL任务路径": etl_task
            })

        # 3. Excel文件生成
        wb = Workbook()
        ws = wb.active

        # 3.1 构建多级表头
        headers = [
            ["序号", "数据来源信息", "", "", "", "数据采集方式", "", "数据进入数仓后的相关信息"],
            ["", "数据来源部门", "数据来源视图名/表名", "数据来源字段名", "来源字段中文含义",
             "采集方式", "ETL任务位置与任务名称", "入仓后的数据表", "入仓后的字段",
             "有无数据", "是否已通过可视化展示", "备注", "数据量", "ETL定时情况"]
        ]

        # 写入表头并设置样式
        bold_font = Font(bold=True)
        for row in headers:
            ws.append(row)
            for cell in ws[ws.max_row]:
                cell.font = bold_font

                # 合并表头区域
        merge_ranges = [
            ('A1:A2', "序号"),
            ('B1:E1', "数据来源信息"),
            ('F1:G1', "数据采集方式"),
            ('H1:N1', "数据入仓信息")
        ]
        for range_str, _ in merge_ranges:
            ws.merge_cells(range_str)

            # 3.2 写入数据内容
        for data_row in dataframe_to_rows(pd.DataFrame(processed_data), index=False, header=False):
            # 列映射逻辑
            ws.append([
                data_row[0],  # A列：序号
                "",  # B列：数据来源部门（空）
                data_row[1],  # C列：数据来源视图名/表名
                data_row[2],  # D列：数据来源字段名
                data_row[3],  # E列：来源字段中文含义
                "",  # F列：采集方式（空）
                data_row[5],  # G列：ETL任务位置与任务名称
                data_row[4],  # H列：入仓后的数据表
                data_row[2],  # I列：入仓后的字段
                *[""] * 5  # J-N列：其他空字段
            ])

        # 3.3 合并相同表名的单元格（C列和G列）
        for table_name, ranges in table_groups.items():
            if ranges['end'] > ranges['start']:
                # 合并C列（数据来源视图名）
                ws.merge_cells(
                    start_row=ranges['start'],
                    end_row=ranges['end'],
                    start_column=3,
                    end_column=3
                )
                # 合并G列（ETL任务位置）
                ws.merge_cells(
                    start_row=ranges['start'],
                    end_row=ranges['end'],
                    start_column=7,
                    end_column=7
                )

        # 3.4 设置列宽（重点加宽G列）
        col_width_config = {
            3: 30,  # C列（数据来源视图名）
            7: 45,  # G列（ETL任务路径）
            8: 25  # H列（入仓表名）
        }
        for col_idx in range(1, 15):
            column_letter = get_column_letter(col_idx)
            # 特殊列宽设置
            if col_idx in col_width_config:
                ws.column_dimensions[column_letter].width = col_width_config[col_idx]
            else:
                ws.column_dimensions[column_letter].width = 18

                # 4. 保存与验证
        wb.save(output_path)
        print(f"✅ 文件生成成功：{output_path}")
        print(f"▸ 合并表名数量：{len(table_groups)} 个")
        print(f"▸ 最新ETL任务路径示例：{processed_data[0]['ETL任务路径']}")

    except Exception as e:
        print(f"❌ 运行错误：{str(e)}")
        if "Permission denied" in str(e):
            print("→ 请关闭已打开的Excel文件后重试")


# 执行示例
if __name__ == "__main__":
    process_excel(
        table2_path="表3.xlsx",
        output_path="数据映射表_20250321.xlsx"
    )